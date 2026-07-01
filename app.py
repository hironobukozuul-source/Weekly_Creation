import streamlit as st
import pandas as pd
import datetime
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import font_manager
from matplotlib.ticker import FuncFormatter
from matplotlib.patches import Rectangle
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from io import BytesIO
import os
import requests

# --- フォント設定 ---
def setup_japanese_font():
    font_path = "NotoSansJP-Regular.ttf"
    if not os.path.exists(font_path):
        url = "https://github.com/googlefonts/noto-cjk/raw/main/Sans/OTF/Japanese/NotoSansCJKjp-Regular.otf"
        r = requests.get(url)
        with open(font_path, 'wb') as f:
            f.write(r.content)
    font_prop = font_manager.FontProperties(fname=font_path)
    font_manager.fontManager.addfont(font_path)
    plt.rcParams['font.family'] = font_prop.get_name()
    return font_prop

jp_font = setup_japanese_font()

# --- 定数・マッピング設定 ---
NAME_MAP = {'Pump': 'Pump', 'Ref1': 'Ref-1', 'Flexible': 'Flexible', 'Ref3': 'Ref-3', 'Ref4': 'Ref-4', 'Ref5': 'Ref-5', 'Awa': 'Awa'}
LINE_START_COLS = {'Pump': 2, 'Ref1': 11, 'Flexible': 20, 'Ref3': 29, 'Ref4': 38, 'Ref5': 47, 'Awa': 56}
DATE_COL = 63
MAINT_KEYWORDS = ['P/C', 'CLN', 'SETUP', '洗浄', 'うがい', 'SPARE', 'C/L', 'QC', '原価改定', '段取', 'メンテナンス', '点検', '清掃', '切替', '予備', 'WAIT', 'SAMPLE']
TOTAL_HOURS = 174 

def to_time(val):
    if isinstance(val, datetime.time): return val
    if isinstance(val, (int, float)):
        ts = int(round(val * 86400))
        return datetime.time((ts // 3600) % 24, (ts // 60) % 60, ts % 60)
    return None

@st.cache_data
def get_available_weeks(df_raw):
    dates = pd.to_datetime(df_raw.iloc[3:, DATE_COL], errors='coerce').dropna()
    mondays = dates[dates.dt.weekday == 0].dt.date.unique()
    return sorted(mondays)

@st.cache_data
def load_sg_cap_mappings(uploaded_file):
    """
    Parses the SG-cap sheet to read the dynamic structural parameters:
    Column D (Index 3) -> SKU/Material Name
    Column G (Index 6) -> Filling Weight in Grams
    """
    try:
        df_cap = pd.read_excel(uploaded_file, sheet_name='SG-cap', header=None)
        mapping = {}
        for r in range(len(df_cap)):
            sku_val = df_cap.iloc[r, 3]    # Column D
            weight_val = df_cap.iloc[r, 6] # Column G
            
            if pd.notna(sku_val) and pd.notna(weight_val):
                sku_str = str(sku_val).strip()
                try:
                    weight_float = float(weight_val)
                    if weight_float > 0:
                        mapping[sku_str] = weight_float
                except ValueError:
                    continue
        return mapping
    except Exception as e:
        st.sidebar.error(f"SG-cap mapping error: {e}")
        return {}

def process_tasks(df_raw):
    tasks = []
    line_config = {}
    for line, start_idx in LINE_START_COLS.items():
        found_ton_col = None
        for c in range(start_idx, start_idx + 10):
            h_vals = [str(df_raw.iloc[r, c]).lower() for r in range(min(3, len(df_raw)))]
            if any('output' in v for v in h_vals) and any('ton' in v for v in h_vals):
                found_ton_col = c
                break
        line_config[line] = {'prod': start_idx, 'start': start_idx+2, 'finish': start_idx+3, 'ton': found_ton_col or (start_idx+4)}
    
    line_states = {}
    for i in range(3, len(df_raw)):
        date_val = df_raw.iloc[i, DATE_COL]
        if not isinstance(date_val, (datetime.datetime, pd.Timestamp)): continue
        base_date = date_val.date()
        for line, cols in line_config.items():
            prod_raw, st_raw, fn_raw, tn_raw = df_raw.iloc[i, cols['prod']], df_raw.iloc[i, cols['start']], df_raw.iloc[i, cols['finish']], df_raw.iloc[i, cols['ton']]
            if pd.isna(st_raw) or pd.isna(fn_raw) or pd.isna(prod_raw): continue
            product = str(prod_raw).strip()
            if product.lower() in ['nan', '連操なし', '']: continue
            s_t, f_t = to_time(st_raw), to_time(fn_raw)
            if s_t and f_t:
                if line not in line_states:
                    line_states[line] = {'current_date': base_date, 'last_start_time': s_t}
                if s_t < line_states[line]['last_start_time']:
                    line_states[line]['current_date'] += datetime.timedelta(days=1)
                if base_date > line_states[line]['current_date']:
                    line_states[line]['current_date'] = base_date
                dt_s = datetime.datetime.combine(line_states[line]['current_date'], s_t)
                dt_f = datetime.datetime.combine(line_states[line]['current_date'], f_t)
                if f_t < s_t: dt_f += datetime.timedelta(days=1)
                try: ton = float(tn_raw) if pd.notna(tn_raw) else 0.0
                except: ton = 0.0
                is_m = (ton <= 0) or any(kw.upper() in product.upper() for kw in MAINT_KEYWORDS)
                tasks.append({'Line': line, 'Product': product, 'Start': dt_s, 'Finish': dt_f, 'Ton': ton, 'is_maint': is_m})
                line_states[line]['last_start_time'] = s_t
    return pd.DataFrame(tasks)

def generate_plot(df_tasks, start_date, sg_cap_map, display_unit):
    plot_start = datetime.datetime.combine(start_date, datetime.time(0, 0))
    plot_end = plot_start + datetime.timedelta(hours=TOTAL_HOURS)
    requested_order = ['Pump', 'Ref1', 'Flexible', 'Ref3', 'Ref4', 'Ref5', 'Awa']
    plot_order = [NAME_MAP[n] for n in requested_order[::-1]]
    line_to_y = {NAME_MAP[n]: i for i, n in enumerate(requested_order[::-1])}

    fig, ax = plt.subplots(figsize=(30, 20), facecolor='white')
    plt.subplots_adjust(top=0.82, bottom=0.08, left=0.08, right=0.95)
    
    box_offset_state = {line: 30 for line in plot_order}
    text_offset_state = {line: 0.05 for line in plot_order}

    merged = []
    for line_key in requested_order:
        line_df = df_tasks[df_tasks['Line'] == line_key].sort_values('Start')
        if line_df.empty: continue
        curr = line_df.iloc[0].to_dict()
        curr['Segments'], curr['TotalTon'] = [(curr['Start'], curr['Finish'])], curr['Ton']
        for idx in range(1, len(line_df)):
            nxt = line_df.iloc[idx].to_dict()
            if nxt['Product'] == curr['Product'] and nxt['is_maint'] == curr['is_maint'] and (nxt['Start'] - curr['Finish']) <= datetime.timedelta(hours=4):
                curr['Finish'], curr['TotalTon'] = nxt['Finish'], curr['TotalTon'] + nxt['Ton']
                curr['Segments'].append((nxt['Start'], nxt['Finish']))
            else:
                merged.append(curr); curr = nxt; curr['Segments'], curr['TotalTon'] = [(curr['Start'], curr['Finish'])], curr['Ton']
        merged.append(curr)

    tick_half_h = 0.05 
    for camp in merged:
        if camp['Finish'] < plot_start or camp['Start'] > plot_end: continue
        line_name = NAME_MAP[camp['Line']]
        y, is_m, color = line_to_y[line_name], camp['is_maint'], ('#7F7F7F' if camp['is_maint'] else '#1F4E78')
        for s_dt, f_dt in camp['Segments']:
            s, e = max(mdates.date2num(s_dt), mdates.date2num(plot_start)), min(mdates.date2num(f_dt), mdates.date2num(plot_end))
            if e > s:
                if is_m: ax.hlines(y, s, e, colors=color, linestyles='dotted', linewidth=4.0, zorder=3)
                else: ax.hlines(y, s, e, colors=color, linewidth=12, capstyle='butt', zorder=3)
                ax.vlines(e, y - tick_half_h, y + tick_half_h, colors=color, linewidth=2.0, zorder=4)
        
        mid = mdates.date2num(max(camp['Start'], plot_start) + (min(camp['Finish'], plot_end) - max(camp['Start'], plot_start))/2)
        if is_m:
            y_text_off = text_offset_state[line_name]
            text_offset_state[line_name] = -0.05 if y_text_off > 0 else 0.05
            ax.text(mid, y + y_text_off, camp['Product'], ha='center', va=('bottom' if y_text_off > 0 else 'top'), fontsize=11, color='#555555', fontweight='bold', fontproperties=jp_font, zorder=5)
        else:
            y_box_off = box_offset_state[line_name]
            box_offset_state[line_name] = -30 if y_box_off > 0 else 30
            
            # Context-Aware Dashboard Metric Switching
            if display_unit == "Pieces (pcs)":
                sku_clean = str(camp['Product']).strip()
                filling_g = sg_cap_map.get(sku_clean, None)
                if filling_g:
                    calculated_pcs = (camp['TotalTon'] * 1000000) / filling_g
                    label_str = f"{camp['Product']}\n{calculated_pcs:,.0f} pcs"
                else:
                    label_str = f"{camp['Product']}\n{camp['TotalTon']:.1f}t\n(No Weight Factor)"
            else:
                label_str = f"{camp['Product']}\n{camp['TotalTon']:.1f}t"
                
            ax.annotate(label_str, xy=(mid, y), xytext=(0, y_box_off), textcoords='offset points', ha='center', va=('bottom' if y_box_off > 0 else 'top'), bbox=dict(boxstyle='square,pad=0.3', fc='white', ec=color, lw=1.5, alpha=0.9), arrowprops=dict(arrowstyle='->', color=color, lw=1), fontsize=11, fontweight='bold', fontproperties=jp_font, zorder=6)

    for y_idx in range(len(plot_order) - 1):
        strip_y = y_idx + 0.5
        ax.axhline(strip_y, color='#F5F5F5', linewidth=20, zorder=1)
        for h_offset in range(0, TOTAL_HOURS, 3):
            t_mark = plot_start + datetime.timedelta(hours=h_offset)
            ax.text(mdates.date2num(t_mark), strip_y, f"{t_mark.hour}", color='#999999', fontsize=10, ha='center', va='center', zorder=2, fontweight='bold')

    ax.set_xlim(mdates.date2num(plot_start), mdates.date2num(plot_end))
    ax.xaxis.set_major_locator(mdates.DayLocator())
    ax.xaxis.set_major_formatter(mdates.DateFormatter('\n%m/%d (%a)'))
    ax.xaxis.set_minor_locator(mdates.HourLocator(byhour=[0, 3, 6, 9, 12, 15, 18, 21]))
    ax.xaxis.set_minor_formatter(FuncFormatter(lambda x, pos: f"{mdates.num2date(x).hour}"))
    ax.set_yticks(range(len(plot_order))); ax.set_yticklabels(plot_order, fontsize=16, fontweight='bold')
    ax.set_ylim(-0.8, 6.8)
    for i in range(9): ax.axvline(mdates.date2num(plot_start + datetime.timedelta(days=i)), color='red', alpha=0.3, linewidth=3, zorder=5)

    ax.text(0.5, 1.12, f"Production Plan - Week of {start_date} (+6hrs)", transform=ax.transAxes, fontsize=48, fontweight='bold', ha='center', va='center', fontproperties=jp_font)
    
    box_w, box_h = 0.033, 0.05
    pos_y = 0.88
    new_pm_x, new_sv_x = 0.883 - 0.011, 0.833 - 0.011 
    fig.patches.append(Rectangle((new_sv_x, pos_y), box_w, box_h, transform=fig.transFigure, fill=False, edgecolor='black', lw=2))
    fig.text(new_sv_x + (box_w/2), pos_y + box_h + 0.005, 'SV', transform=fig.transFigure, ha='center', fontweight='bold', fontsize=16)
    fig.patches.append(Rectangle((new_pm_x, pos_y), box_w, box_h, transform=fig.transFigure, fill=False, edgecolor='black', lw=2))
    fig.text(new_pm_x + (box_w/2), pos_y + box_h + 0.005, 'PM', transform=fig.transFigure, ha='center', fontweight='bold', fontsize=16)

    buf = BytesIO(); plt.savefig(buf, format='png'); plt.close(); buf.seek(0)
    return buf

# --- UI / メインロジック ---
st.set_page_config(layout="wide", page_title="Weekly Production Master Report")
st.title("🏭 Weekly Production Master Report Generator")

# Interactive UI Control Mode toggle for runtime chart rendering switching
display_unit = st.sidebar.radio("Dashboard Gantt Display Unit Mode:", ["Tonnage (t)", "Pieces (pcs)"])

uploaded_file = st.file_uploader("Excelファイルをアップロード (.xlsm)", type=["xlsm"])

if uploaded_file:
    df_raw = pd.read_excel(uploaded_file, sheet_name='Fill', header=None)
    
    # Instantiate lookups natively
    sg_cap_map = load_sg_cap_mappings(uploaded_file)
    if sg_cap_map:
        st.sidebar.success(f"Successfully loaded {len(sg_cap_map)} product configurations from 'SG-cap'!")
    else:
        st.sidebar.warning("No mappings loaded. Check 'SG-cap' layout column names.")
        
    available_weeks = get_available_weeks(df_raw)
    if available_weeks:
        selected_week = st.selectbox("対象週を選択", available_weeks)
        if st.button("🚀 レポート生成開始"):
            with st.spinner('計算と描画を行っています...'):
                df_tasks = process_tasks(df_raw)
                img_buf = generate_plot(df_tasks, selected_week, sg_cap_map, display_unit)
                
                # --- Excel生成 ---
                wb = openpyxl.Workbook()
                thick_black = Side(style='thick', color='000000')
                start_dt_excel = datetime.datetime.combine(selected_week, datetime.time(0, 0))
                hour_list = [start_dt_excel + datetime.timedelta(hours=h) for h in range(TOTAL_HOURS)]
                
                # 1. Hourly_Volume シート (Tonnage Engine)
                ws_vol = wb.active
                ws_vol.title = "Hourly_Volume"
                
                ws_vol.cell(1, 1, "Line").font = Font(bold=True)
                ws_vol.cell(1, 2, "Product").font = Font(bold=True)
                for i, h_dt in enumerate(hour_list):
                    next_h = h_dt + datetime.timedelta(hours=1)
                    header_str = f"{h_dt.strftime('%m/%d %H:%M')}~{next_h.strftime('%H:%M')}"
                    cell = ws_vol.cell(1, 3 + i, header_str)
                    cell.font = Font(bold=True); cell.alignment = Alignment(text_rotation=90, horizontal='center')

                clean_tasks = df_tasks[~df_tasks['is_maint']]
                unique_items = sorted(list(set(zip(clean_tasks['Line'], clean_tasks['Product']))))
                for r_idx, (line, product) in enumerate(unique_items):
                    row_num = r_idx + 2
                    ws_vol.cell(row_num, 1, line); ws_vol.cell(row_num, 2, product)
                    for c_idx, h_dt in enumerate(hour_list):
                        h_end = h_dt + datetime.timedelta(hours=1)
                        overlap_tasks = clean_tasks[(clean_tasks['Line'] == line) & (clean_tasks['Product'] == product)]
                        ton_sum = sum(t['Ton'] * ((min(t['Finish'], h_end) - max(t['Start'], h_dt)).total_seconds() / (t['Finish'] - t['Start']).total_seconds()) for _, t in overlap_tasks.iterrows() if (min(t['Finish'], h_end) - max(t['Start'], h_dt)).total_seconds() > 0)
                        if ton_sum > 0:
                            cell = ws_vol.cell(row_num, 3 + c_idx, round(ton_sum, 2))
                            cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

                for col_idx in range(3, ws_vol.max_column + 1):
                    header = str(ws_vol.cell(1, col_idx).value)
                    if "00:00" in header:
                        for r in range(1, ws_vol.max_row + 1): ws_vol.cell(r, col_idx).border = Border(left=thick_black)

                # =========================================================
                # NEW ADDITION: 2. Hourly_Pieces Sheet Generation Engine
                # =========================================================
                ws_pcs = wb.create_sheet("Hourly_Pieces")
                ws_pcs.cell(1, 1, "Line").font = Font(bold=True)
                ws_pcs.cell(1, 2, "Product").font = Font(bold=True)
                
                # Format time headers identical to Hourly_Volume
                for i, h_dt in enumerate(hour_list):
                    next_h = h_dt + datetime.timedelta(hours=1)
                    header_str = f"{h_dt.strftime('%m/%d %H:%M')}~{next_h.strftime('%H:%M')}"
                    cell = ws_pcs.cell(1, 3 + i, header_str)
                    cell.font = Font(bold=True); cell.alignment = Alignment(text_rotation=90, horizontal='center')

                # Calculate matrix intersections for piece targets
                for r_idx, (line, product) in enumerate(unique_items):
                    row_num = r_idx + 2
                    ws_pcs.cell(row_num, 1, line); ws_pcs.cell(row_num, 2, product)
                    
                    # Fetch filling specification from SG-cap matrix
                    sku_name_clean = str(product).strip()
                    filling_grams = sg_cap_map.get(sku_name_clean, None)
                    
                    for c_idx, h_dt in enumerate(hour_list):
                        h_end = h_dt + datetime.timedelta(hours=1)
                        overlap_tasks = clean_tasks[(clean_tasks['Line'] == line) & (clean_tasks['Product'] == product)]
                        ton_sum = sum(t['Ton'] * ((min(t['Finish'], h_end) - max(t['Start'], h_dt)).total_seconds() / (t['Finish'] - t['Start']).total_seconds()) for _, t in overlap_tasks.iterrows() if (min(t['Finish'], h_end) - max(t['Start'], h_dt)).total_seconds() > 0)
                        
                        if ton_sum > 0:
                            if filling_grams:
                                # Tonnage * 1M grams / item unit grams
                                total_pcs_calculated = (ton_sum * 1000000) / filling_grams
                                cell = ws_pcs.cell(row_num, 3 + c_idx, int(round(total_pcs_calculated)))
                            else:
                                # Traceable error output if SKU configuration data is missing
                                cell = ws_pcs.cell(row_num, 3 + c_idx, "Missing SG-cap Specs")
                            
                            # Soft blue highlight variant for structural distinction from volume page
                            cell.fill = PatternFill(start_color="C9DAF8", end_color="C9DAF8", fill_type="solid")

                # Apply midnight vertical break rule boundaries 
                for col_idx in range(3, ws_pcs.max_column + 1):
                    header = str(ws_pcs.cell(1, col_idx).value)
                    if "00:00" in header:
                        for r in range(1, ws_pcs.max_row + 1): ws_pcs.cell(r, col_idx).border = Border(left=thick_black)
                # =========================================================

                # 3. Visual_Schedule シート
                ws_vis = wb.create_sheet("Visual_Schedule")
                img_for_excel = BytesIO(img_buf.getvalue())
                ws_vis.add_image(openpyxl.drawing.image.Image(img_for_excel), 'B2')
                
                out_excel = BytesIO()
                wb.save(out_excel)
                
                # --- 結果表示 ---
                st.image(img_buf, use_container_width=True)
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button("📥 Excelレポートを保存", out_excel.getvalue(), f"Production_Report_{selected_week}.xlsx")
                with col2:
                    st.download_button("🖼️ ガントチャート画像を保存", img_buf.getvalue(), f"Gantt_{selected_week}.png")
